import boto3
from kivy.app import App
from kivy.core.window import Window
from kivy.graphics import Color, RoundedRectangle
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from openpyxl import load_workbook
from io import BytesIO

from MyButton import MyButton
from S3_access_key import S3_key as Key
from kivy.uix.screenmanager import Screen


class MyFlashcardsScreen(Screen):

    def on_pre_enter(self):
        self.load_folders()
        self.start_video()

    def load_folders(self):
        app = App.get_running_app()
        user_login = app.login
        s3 = boto3.client(
            's3',
            aws_access_key_id=Key.aws_access_key_id,
            aws_secret_access_key=Key.aws_secret_access_key,
            region_name=Key.region_name
        )
        bucket_name = 'fank'
        prefix = f"{user_login}/"
        result = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix, Delimiter='/')

        folders = result.get('CommonPrefixes')
        self.ids.folders_box.clear_widgets()
        if folders:
            for folder in folders:
                folder_name = folder.get('Prefix')[len(prefix):-1]
                btn = MyButton(text=folder_name, size_hint_y=None, height=Window.height * 0.09,
                             font_size=Window.height * 0.04, background_color=(0.678, 0.847, 0.902, 0.6))
                btn.bind(on_release=lambda x, folder=folder_name: self.show_flashcards(folder))
                self.ids.folders_box.add_widget(btn)
        else:
            self.show_popup("Informacja", "Nie ma takiego folderu.")

    def show_flashcards(self, folder_name):
        app = App.get_running_app()
        user_login = app.login
        s3 = boto3.client(
            's3',
            aws_access_key_id=Key.aws_access_key_id,
            aws_secret_access_key=Key.aws_secret_access_key,
            region_name=Key.region_name
        )
        bucket_name = 'fank'
        prefix = f"{user_login}/{folder_name}/"
        result = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

        flashcards = []
        if 'Contents' in result:
            for file in result['Contents']:
                if file['Key'].endswith('.xlsx'):
                    response = s3.get_object(Bucket=bucket_name, Key=file['Key'])
                    file_content = response['Body'].read()
                    wb = load_workbook(filename=BytesIO(file_content))
                    ws = wb.active
                    for row in ws.iter_rows(values_only=True):
                        flashcard = {
                            'front': row[0],
                            'back': row[1],
                            'progress': row[2] if len(row) > 2 else 0
                        }
                        flashcards.append(flashcard)

        if flashcards:
            learning_screen = self.manager.get_screen('learning_screen')
            learning_screen.flashcards = flashcards
            self.manager.current = 'learning_screen'
        else:
            self.show_popup("Informacja", "W folderze nie ma żadnych fiszek")

    def display_cards(self, cards_content):
        content_label = Label(text=cards_content, size_hint_y=None, height=400)
        scroll_view = ScrollView(size_hint=(1, None), size=(400, 400))
        scroll_view.add_widget(content_label)

        box_layout = BoxLayout(orientation='vertical')
        box_layout.add_widget(scroll_view)
        close_button = Button(text="Close", size_hint=(1, None), height=50)
        close_button.bind(on_release=lambda x: popup.dismiss())

        box_layout.add_widget(close_button)
        popup = Popup(title="Flashcard Details", content=box_layout, size_hint=(None, None), size=(420, 500))
        popup.open()

    def show_popup(self, title, message):
        content = BoxLayout(orientation='vertical', spacing=5, padding=5)
        content.add_widget(Label(text=message))
        close_btn = Button(text='OK', size_hint=(1, 0.25))
        content.add_widget(close_btn)

        popup = Popup(title=title, content=content, size_hint=(None, None), size=(400, 200))

        close_btn.bind(on_press=popup.dismiss)

        popup.open()

    def on_leave(self):
        self.stop_video()

    def start_video(self):
        self.ids.video.state = 'play'

    def stop_video(self):
        self.ids.video.state = 'stop'
