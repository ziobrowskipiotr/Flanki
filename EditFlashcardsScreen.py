import os
import tempfile
import subprocess
from random import seed

import boto3
import s3fs
from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.screenmanager import Screen
from kivy.core.window import Window
from kivy.uix.textinput import TextInput
from openpyxl import load_workbook
from io import BytesIO

from RoundedTextInput import RoundedTextInput
from S3_access_key import S3_key as Key

class EditFlashcardsScreen(Screen):

    def on_pre_enter(self):
        self.load_folders()
        self.start_video()

    def load_folders(self):
        app = App.get_running_app()
        user_login = app.login
        s3 = boto3.client(
            's3',
            aws_access_key_id=Key.aws_access_key_id,
            aws_secret_access_key=Key.aws_secret_access_key,
            region_name=Key.region_name
        )
        bucket_name = 'fank'
        prefix = f"{user_login}/"
        result = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix, Delimiter='/')

        folders = result.get('CommonPrefixes')
        self.ids.folders_box.clear_widgets()
        if folders:
            for folder in folders:
                folder_name = folder.get('Prefix')[len(prefix):-1]
                btn = Button(text=folder_name, size_hint_y=None, height=Window.height * 0.09, font_size=Window.height * 0.04, background_color= (0.678, 0.847, 0.902, 0.6))
                btn.bind(on_release=lambda x, folder=folder_name: self.show_flashcards(folder))
                self.ids.folders_box.add_widget(btn)
        else:
            self.show_popup("Informacja", "Nie ma takiego folderu.")

    def show_flashcards(self, folder_name):
        app = App.get_running_app()
        user_login = app.login
        s3 = boto3.client(
            's3',
            aws_access_key_id=Key.aws_access_key_id,
            aws_secret_access_key=Key.aws_secret_access_key,
            region_name=Key.region_name
        )
        bucket_name = 'fank'
        prefix = f"{user_login}/{folder_name}/"
        result = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)

        flashcards = [obj['Key'] for obj in result.get('Contents', []) if obj['Key'].endswith('.xlsx')]

        self.ids.files_box.clear_widgets()
        for file_key in sorted(flashcards):
            base_name = os.path.basename(file_key)
            file_index = base_name.split('.')[0]
            formatted_name = f"{file_index}.xlsx"
            seed()
            btn = Button(text=formatted_name, size_hint_y=None, height=Window.height * 0.07, font_size=Window.height * 0.04, background_color= (0.678, 0.847, 0.902, 0.6))
            btn.bind(on_release=lambda x, fk=f"s3://{bucket_name}/{file_key}": self.edit_flashcard(fk))
            self.ids.files_box.add_widget(btn)

    def upload_to_s3(self, local_path, s3_path):
        s3 = s3fs.S3FileSystem(anon=False,
                               key=Key.aws_access_key_id,
                               secret=Key.aws_secret_access_key,
                               client_kwargs={'region_name': Key.region_name})

        with open(local_path, 'rb') as f:
            with s3.open(s3_path, 'wb') as s3_file:
                s3_file.write(f.read())

        print(f"Plik {os.path.basename(s3_path)} został przesłany z powrotem do S3.")

    def download_excel(self, s3_path):
        s3 = s3fs.S3FileSystem(anon=False,
                               key=Key.aws_access_key_id,
                               secret=Key.aws_secret_access_key,
                               client_kwargs={'region_name': Key.region_name})

        local_path = os.path.join(tempfile.gettempdir(), os.path.basename(s3_path))
        with s3.open(s3_path, 'rb') as f:
            with open(local_path, 'wb') as local_file:
                local_file.write(f.read())

        return local_path

    def save_changes(self):
        if hasattr(self, 'current_local_path') and hasattr(self, 'current_s3_path'):
            self.upload_to_s3(self.current_local_path, self.current_s3_path)
            self.clear_flashcards_view()
            self.show_popup("Udane", "Fiszka została edytowana")
        else:
            self.show_popup("Błąd", "Coś poszło nie tak.")

    def clear_flashcards_view(self):
        self.ids.files_box.clear_widgets()
        self.load_folders()

    def show_popup(self, title, message):
        content = BoxLayout(orientation='vertical', spacing=5, padding=5)
        content.add_widget(Label(text=message))
        close_btn = Button(text='OK', size_hint=(1, 0.25))
        content.add_widget(close_btn)

        # Używamy Window.size, aby obliczyć rozmiary względne do okna aplikacji
        window_width, window_height = Window.size
        popup_width = window_width * 0.8  # 80% szerokości okna
        popup_height = window_height * 0.5  # 50% wysokości okna

        popup = Popup(title=title,
                      content=content,
                      size_hint=(None, None),
                      size=(popup_width, popup_height))

        close_btn.bind(on_press=popup.dismiss)

        popup.open()

    def edit_flashcard(self, s3_path):
        local_path = self.download_excel(s3_path)  # Zmieniłem nazwę dla klarowności
        self.current_local_path = local_path
        self.current_s3_path = s3_path
        self.load_excel_for_editing(local_path)  # Ładowanie danych do edycji w aplikacji

    def load_excel_for_editing(self, path):
        workbook = load_workbook(path)
        sheet = workbook.active
        data_to_edit = {'cell1': sheet['A1'].value, 'cell2': sheet['B1'].value}

        # Now you can display these values in Kivy TextInput widgets, allow users to edit, and save them back
        self.show_edit_popup(data_to_edit, path)

    def show_edit_popup(self, data, path):
        content = BoxLayout(orientation='vertical', spacing=5, padding=5)
        input1 = RoundedTextInput(text=str(data['cell1']), multiline=False, hint_text_color=[1,1,1,1], foreground_color=[1,1,1,1], background_color=[0,0,0,0], background_image= "", background_normal= "", background_active= "", size_hint= (0.8 ,0.12), pos_hint= {'center_x':0.5, 'center_y':0.5}, )
        input2 = RoundedTextInput(text=str(data['cell2']), multiline=False, hint_text_color=[1,1,1,1], foreground_color=[1,1,1,1], background_color=[0,0,0,0], background_image= "", background_normal= "", background_active= "", size_hint= (0.8 ,0.12), pos_hint= {'center_x':0.5, 'center_y':0.65}, )
        content.add_widget(input1)
        content.add_widget(input2)

        save_btn = Button(text='Save', size_hint=(1, 0.12))
        content.add_widget(save_btn)

        # Tutaj ustawiamy size_hint na pewien procent aktualnej szerokości i wysokości okna

        popup = Popup(title="Edit Cells", content=content, size_hint=(0.8, 0.34))
        save_btn.bind(on_press=lambda x: self.save_edited_data(path, input1.text, input2.text, popup))
        popup.open()

    def save_edited_data(self, path, value1, value2, popup):
        workbook = load_workbook(path)
        sheet = workbook.active
        sheet['A1'] = value1
        sheet['B1'] = value2
        workbook.save(path)
        popup.dismiss()
        self.upload_to_s3(path, self.current_s3_path)  # Automatically save back to S3
        self.show_popup("Success", "Changes saved successfully")

    def on_leave(self):
        self.stop_video()

    def start_video(self):
        self.ids.video.state = 'play'

    def stop_video(self):
        self.ids.video.state = 'stop'

    def create(self):
        self.create_new_flashcard_folder()
        app = App.get_running_app()
        app.generate_excel()
        os.remove(app.file_name)
